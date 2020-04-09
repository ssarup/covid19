from datetime import datetime
import openpyxl
from covid19io.OutputWriter import OutputWriter


class ExcelWriter(OutputWriter):
    def __init__(self, filename_, openExisting=False, hasMacros=False):
        print('xx', filename_)
        self._filename = filename_
        if openExisting:
            self._workbook = openpyxl.load_workbook(filename=self._filename, keep_vba=hasMacros)
        else:
            self._workbook = openpyxl.Workbook()
        self._worksheet = self._workbook.active
        self._activeSheetName = self._worksheet.title
        self._rowNum = 1
        self._column = 1

    @property
    def filename(self):
        return self._filename

    @property
    def activeSheetName(self):
        return self._activeSheetName

    @staticmethod
    def _setFormatting(cell_, valueForFormatting_):
        if type(valueForFormatting_) == int:
            cell_.number_format = openpyxl.styles.numbers.FORMAT_NUMBER
        elif type(valueForFormatting_) == datetime:
            cell_.number_format = 'MM/DD/YYYY'
        elif type(valueForFormatting_) == str:
            cell_.number_format = openpyxl.styles.numbers.FORMAT_TEXT

    def setActiveSheet(self, sheetName_):
        self._worksheet = self._workbook.create_sheet(sheetName_)
        self._activeSheetName = sheetName_
        self._rowNum = 1
        self._column = 1

    def writeLine(self, tup_, header_):
        assert isinstance(tup_, tuple)
        assert isinstance(header_, bool)
        for item in tup_:
            currCell = self._worksheet.cell(row=self._rowNum, column=self._column)
            # print('xx', type(item), item, str(tup_))
            ExcelWriter._setFormatting(currCell, item)
            currCell.value = item
            self._column = self._column + 1
        self._rowNum = self._rowNum + 1
        self._column = 1

    def save(self, filename=None):
        if filename is None:
            self._workbook.save(filename=self._filename)
        else:
            self._workbook.save(filename=filename)

